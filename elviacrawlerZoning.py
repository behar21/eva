#!/usr/bin/env python
# -*- coding: utf-8 -*-
import webbrowser
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.keys import Keys
import openpyxl
from tqdm import tqdm
import time
#--------------- Start Excel -----------------#
wb = openpyxl.Workbook()
wb = openpyxl.load_workbook(filename='ElviaZoningSample.xlsx')
sheets = wb.sheetnames
ws = wb[sheets[0]]
#--------------- End Excel -------------------#
pbar = tqdm(total=50)

#----------------- Read Excel Rows -----------------#    	
for i in range(201,301):
	try:
    	
		index = str(i)
		#--------- Read Excel Data -----------------#
		brand = ws['P'+index].value
		model = ws['T'+index].value+" "+ws['U'+index].value
		inverhrkersetzung = ws['V'+index].value
		leasing = ws['H'+index].value
		accessories = ws['R'+index].value
		fahrzeugverwendung = ws['I'+index].value
		gender = ws['J'+index].value
 		licenceAge = ws['K'+index].value
		zip = ws['M'+index].value
		nationality = ws['L'+index].value
		milage = ws['G'+index].value
		garage = ws['F'+index].value
		deducible = ws['O'+index].value
       		educibleVol = ws['P'+index].value
        	kanton = ws['N'+index].value
		chrome = webdriver.Chrome()
		chrome.get("https://tarif.elvia.ch/direct/start.do?calc=mfz&variante=pw&lang=de")
		prit = WebDriverWait(chrome,10)
  		
		#--------- 1st Page -----------------------#
        	element = prit.until(EC.element_to_be_clickable((By.ID, 'id6731')))
		dp1=chrome.find_element_by_xpath("//*[@id='id6731']/option[text()='2017']").click()
		
		element = prit.until(EC.element_to_be_clickable((By.ID, '8393')))
		dp2=chrome.find_element_by_xpath("//*[@id='8393']/option[text()='AUDI']").click()
		
	
		element = prit.until(EC.element_to_be_clickable((By.ID, '8440')))
		dp3=chrome.find_element_by_xpath("//*[@id='8440']/option[text()='Q5/SQ5 Modell 2016-']").click()
		
		element = prit.until(EC.element_to_be_clickable((By.ID, 'id6815')))
		dp4=chrome.find_element_by_xpath("//*[@id='id6815']/option[text()='Q5 2.0 TDI design quattro']").click()
		
		
			
    		#chrome.find_element_by_xpath("//*[@id='8572']").click()
    			
  		element = prit.until(EC.element_to_be_clickable((By.ID,'1730')))
  		if chrome.find_element_by_xpath("//*[@id='8834']"):
				
  			table_rows = chrome.find_elements_by_id("8834")
  			for row in table_rows:
    				chrome.execute_script("arguments[0].style.display = 'block'; return arguments[0];", row)
  				code = row.find_element_by_id("8907")
  				if code.get_attribute('innerHTML') == "01AF225":
  					button = row.find_element_by_id("1730")
  					button.click()
  					break

  		
  		
  		#-------Go to 2nd Page -------------

    		element = prit.until(EC.element_to_be_clickable((By.ID,'id8214')))
    		chrome.find_element_by_xpath("//*[@id='id8214']/option[text()='"+str(kanton)+"']").click()
   		#kontrollschild = chrome.find_element_by_xpath("//*[@id='id9108']") 
   		#kontrollschild.clear()
   		#kontrollschild.send_keys("123456")
   	
 		inverkehrestzung = chrome.find_element_by_xpath("//*[@id='id9116']")
   	 	inverkehrestzung.clear()
   	 	inverkehrestzung.send_keys("01.01.2017")
  			
		if leasing == "yes":	
  	  		chrome.find_element_by_xpath("//*[@id='id2640']/option[text()='Leasing']").click()
		else:
	  		chrome.find_element_by_xpath("//*[@id='id2640']/option[text()='kein Leasing']").click()	
    	
		if fahrzeugverwendung == 'Private':
			chrome.find_element_by_xpath("//*[@id='id1000556']/option[text()='Privat']").click()
		else:
			chrome.find_element_by_xpath("//*[@id='id1000556']/option[text()='Privat und geschäftlich']").click()
		
		#------- Got  to 3rd Page ------------
 	   	chrome.find_element_by_xpath("//*[@id='1730']").click()	
		
		element = prit.until(EC.element_to_be_clickable((By.ID,'23729')))
		
		if gender == 'Male':
	        	
			chrome.find_element_by_id("23738").click()
		else:
            		
			chrome.find_element_by_id("23732").click()
		
  		element = prit.until(EC.element_to_be_clickable((By.ID,'id10558')))
    		lAge = '01.01.'+str(licenceAge)
		fuhrerausweis = chrome.find_element_by_xpath("//*[@id='id10558']")
    		fuhrerausweis.clear()
    		fuhrerausweis.send_keys(lAge)
	
	  	plz = chrome.find_element_by_xpath("//*[@id='id1001001']")
	  	plz.clear()
	   	plz.send_keys(str(zip)) 
	    	chrome.find_element_by_xpath("//*[@id='id10558']").click()
        
        	if chrome.find_elements_by_xpath("//*[@id='id1004857']"):
            		chrome.find_element_by_xpath("//*[@id='id1004857']/option[2]").click()
    	        #       ort.send_keys(RETURN)
            	#	ort.send_keys(ARROW_DOWN)
            	#	ort.send_keys(RETURN)

		
		#---- Milage -----------------
		element = prit.until(EC.element_to_be_clickable((By.ID,'id1744')))
	    	jahr = chrome.find_element_by_xpath("//*[@id='id1744']")
	    	jahr.clear()
	    	milageJahr = '5'
		if milage  == '0-2999':
			milageJahr = '2000'
		elif milage == '3000-5999':
			milageJahr = '5000'
		elif milage == '6000-9999':
			milageJahr = '8000'
		elif milage == '10000-14999':
			milageJahr = '12500'
		elif milage == '15000-24999':
			milageJahr = '20000'
		elif milage == '25000+':
			milageJahr = '30000'
		jahr.send_keys(milageJahr)
	    	

		birthdate = chrome.find_element_by_xpath("//*[@id='8120']")
	    	birthdate.clear()
		
		if str(licenceAge) == "2013":
			birthdate.send_keys('01.01.1994')
		elif str(licenceAge) == "1996":
                        birthdate.send_keys('01.01.1976')
		else:
	    		birthdate.send_keys('01.01.1996')
	    	
		#--nationality---
		#chrome.find_element_by_xpath("//*[@id='id3329']/option[text() ='"+nationality+"']").click()
		
		#if nationality != 'Schweiz':
		#	chrome.find_element_by_xpath("//*[@id='id3348']/option[text()='C (Niederlassung)']").click()
		
		#----  Wird Ihr Fahrzeug von Lenkern unter 25 Jahren gefahren?  -----
		chrome.find_element_by_xpath("//*[@id='id11006503']").click()
		
		#---- Garage ----------------
		if garage == 'None':
			chrome.find_element_by_xpath("//*[@id='id11006431']").click()
		else:
			chrome.find_element_by_xpath("//*[@id='id11006417']").click()
	
		chrome.find_element_by_xpath("//*[@id='id3329']/option[text() ='"+nationality+"']").click()

		if nationality != 'Schweiz':
                        chrome.find_element_by_xpath("//*[@id='id3348']/option[text()='C (Niederlassung)']").click()
		
		if str(licenceAge) == "1996":
			chrome.find_element_by_xpath("//*[@id='id1001369']").click()
		
		
		#---------------_Last Page ---------------------------------
		chrome.find_element_by_xpath("//*[@id='id10558']").click()
	    	chrome.find_element_by_xpath("//*[@id='1730']").click()
		
	   	element = prit.until(EC.element_to_be_clickable((By.ID,'1730')))
	    	chrome.find_element_by_xpath("//*[@id='1730']").click()
	    	
		
		#-------Get Heftplicht -------
		#------delselect all-----------
        	if leasing == "yes":
            		chrome.find_element_by_name("I9.checked").click()
        
        	else :
            		chrome.find_element_by_name("I10.checked").click()
		
	
		element = prit.until(EC.element_to_be_clickable((By.ID,'id4861')))
	 	haft = chrome.find_element_by_xpath("//*[@id='id100071']")
		#---Write Excel-------
		ws['X'+index] = haft.text
		
		#------Check Teilkasko---------
        	element = prit.until(EC.element_to_be_clickable((By.NAME,'I10.checked')))
                chrome.find_element_by_name("I10.checked").click()
		
		#------Uncheck Migefuhrte-----------------
     			
		element = prit.until(EC.element_to_be_clickable((By.NAME,'I14.checked')))
		chrome.find_element_by_name("I14.checked").click()
        
		element = prit.until(EC.element_to_be_clickable((By.NAME,'I12.sval')))
            	#if str(deducible) == "300":
               	chrome.find_element_by_xpath("//*[@name='I12.sval']/option[text()='300.00']").click()
                #else:
                #chrome.find_element_by_xpath("//*[@name='I12.sval']/option[text()='0.00']").click()
		
		element = prit.until(EC.element_to_be_clickable((By.ID,'id100071')))						
    		teil =  chrome.find_element_by_xpath("//*[@id='id100071']")
		#---Write Excel------
	
		
		ws['Y'+index] = teil.text
		#------Select Vollkaso checkbox----------
		element = prit.until(EC.element_to_be_clickable((By.NAME,'I9.checked')))
        	chrome.find_element_by_name("I9.checked").click()
		
		chrome.find_element_by_xpath("//*[@name='I12.sval']/option[text()='500.00']").click()
     	        #if str(deducibleVol) == "300":
            	chrome.find_element_by_xpath("//*[@name='I14.sval']/option[text()='300.00']").click()
            	
	
		element = prit.until(EC.element_to_be_clickable((By.NAME,'I16.checked')))
		chrome.find_element_by_name("I16.checked").click()	
		element = prit.until(EC.element_to_be_clickable((By.ID,'id100071')))  
		vol = chrome.find_element_by_xpath("//*[@id='id100071']")
        	#---Write Excel------
		ws['Z'+index] =vol.text
 
	except :
		wb.save("ElviaZoningSample.xlsx")
		outY = open("Log.txt","a")
		outY.write("Error at ID: "+index)
		outY.close()
		#chrome.close()
		#chrome.quit()
		pass
	finally :
		pbar.update(1)
		chrome.close()
		chrome.quit()
wb.save("ElviaZoningSample.xlsx")
pbar.close()
